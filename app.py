import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# 페이지 설정
st.set_page_config(page_title="연도별 규격 집계", layout="wide")

st.title("📅 연도별 & 규격별 출고 집계")
st.markdown("""
파일을 업로드하면 **연도(Year)** 별로 묶고, 그 안에서 **규격**별로 합쳐서 
**총 수량(갯수)**과 **총 합계금액**을 보여줍니다.
""")

# 1. 파일 업로드
uploaded_file = st.file_uploader("엑셀 또는 CSV 파일을 업로드하세요", type=['xlsx', 'csv'])

if uploaded_file is not None:
    st.info("연도별 데이터를 분석 중입니다...")

    # 2. 데이터 읽기
    @st.cache_data
    def load_data(file):
        file.seek(0)
        try:
            return pd.read_excel(file, header=2), "Excel"
        except:
            pass
        encodings = ['utf-8', 'cp949', 'euc-kr']
        for enc in encodings:
            try:
                file.seek(0)
                return pd.read_csv(file, header=2, encoding=enc), f"CSV({enc})"
            except:
                pass
        return None, "Fail"

    df, msg = load_data(uploaded_file)

    if df is not None:
        try:
            # 3. 전처리
            df.columns = df.columns.astype(str).str.strip()
            if '규 격' in df.columns:
                df = df[~df['규 격'].astype(str).str.contains('합계', na=False)]

            date_col = '납품일'
            spec_col = '규 격'
            qty_col = '수량'
            price_col = '합계금액'
            unit_col = '단위'

            if date_col in df.columns and spec_col in df.columns:
                df[date_col] = df[date_col].ffill()
                df[spec_col] = df[spec_col].fillna("규격 미기재")
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                
                # [핵심 변경] 연도 추출 (예: 2024년, 2025년)
                df['연도'] = df[date_col].dt.strftime('%Y년')

                # 숫자 변환
                for col in [qty_col, price_col]:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.replace(',', '', regex=False)
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

                # 4. 그룹화 (연도 -> 규격)
                agg_dict = {}
                if qty_col in df.columns: agg_dict[qty_col] = 'sum'
                if price_col in df.columns: agg_dict[price_col] = 'sum'
                if unit_col in df.columns: agg_dict[unit_col] = 'first'

                df_grouped = df.groupby(['연도', spec_col]).agg(agg_dict)

                # 컬럼 순서
                cols_order = []
                if unit_col in df_grouped.columns: cols_order.append(unit_col)
                if qty_col in df_grouped.columns: cols_order.append(qty_col)
                if price_col in df_grouped.columns: cols_order.append(price_col)
                
                df_final = df_grouped[cols_order]

                st.success(f"집계 완료! 총 {len(df_final)}개의 규격 항목이 정리되었습니다.")
                st.dataframe(df_final)

                # 5. 엑셀 저장 및 스타일링
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_final.to_excel(writer, sheet_name='연도별집계')

                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active

                # 스타일
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # 연한 주황색
                center_align = Alignment(horizontal='center', vertical='center')
                right_align = Alignment(horizontal='right', vertical='center')

                # 숫자 컬럼 인덱스 찾기
                number_col_indices = []
                for cell in ws[1]:
                    if cell.value in [qty_col, price_col, '금액']:
                        number_col_indices.append(cell.column)

                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.fill = header_fill
                            cell.font = Font(bold=True)
                            cell.alignment = center_align
                        else:
                            if cell.column in number_col_indices:
                                cell.alignment = right_align
                                cell.number_format = '#,##0'
                            else:
                                cell.alignment = center_align

                # 너비 조정
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="📥 연도별 집계 파일 다운로드",
                    data=output,
                    file_name="연도별_규격별_집계표.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                st.error("필수 컬럼을 찾을 수 없습니다.")
        except Exception as e:
            st.error(f"오류 발생: {e}")
    else:
        st.error("파일 형식 오류")
